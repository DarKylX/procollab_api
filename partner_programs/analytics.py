import io
from decimal import Decimal, ROUND_HALF_UP

from django.conf import settings
from django.db.models import Count, Prefetch
from django.utils import timezone
from openpyxl import Workbook

from core.utils import sanitize_excel_value
from partner_programs.models import (
    PartnerProgram,
    PartnerProgramProject,
    PartnerProgramUserProfile,
)
from projects.models import Collaborator, ProjectLink
from project_rates.models import (
    ProjectEvaluation,
    ProjectEvaluationScore,
    ProjectExpertAssignment,
)


EVALUATION_STATUS_NOT_EVALUATED = "not_evaluated"
EVALUATION_STATUS_PARTIALLY_EVALUATED = "partially_evaluated"
EVALUATION_STATUS_EVALUATED = "evaluated"

SUBMISSION_STATUS_NOT_SUBMITTED = "not_submitted"
SUBMISSION_STATUS_SUBMITTED = "submitted"


def build_program_analytics_payload(program: PartnerProgram) -> dict:
    links = list(_get_program_links(program))
    assignment_counts = _get_assignment_counts(program, links)
    default_required = _get_default_required_evaluations(program)

    submissions = [
        _build_submission(
            program_project=program_project,
            assignment_counts=assignment_counts,
            default_required=default_required,
        )
        for program_project in links
    ]

    submitted_projects_count = sum(
        1
        for submission in submissions
        if submission["submission_status"] == SUBMISSION_STATUS_SUBMITTED
    )
    evaluated_projects_count = sum(
        1
        for submission in submissions
        if submission["submission_status"] == SUBMISSION_STATUS_SUBMITTED
        and submission["evaluation_status"] == EVALUATION_STATUS_EVALUATED
    )
    project_average_scores = [
        Decimal(str(submission["average_score"]))
        for submission in submissions
        if submission["average_score"] is not None
    ]

    return {
        "program_id": program.id,
        "title": program.name,
        "status": program.status,
        "current_stage": _get_current_stage(program),
        "evaluation_deadline": program.datetime_evaluation_ends,
        "participants_count": _get_registered_participants_count(program),
        "submitted_projects_count": submitted_projects_count,
        "evaluated_projects_count": evaluated_projects_count,
        "average_score": _decimal_to_number(_average(project_average_scores)),
        "submissions": submissions,
    }


def build_program_analytics_xlsx(
    program: PartnerProgram,
    *,
    include_contacts: bool = False,
) -> bytes:
    payload = build_program_analytics_payload(program)
    submissions = payload["submissions"]
    program._privacy_include_contacts = include_contacts

    workbook = Workbook()
    projects_sheet = workbook.active
    projects_sheet.title = "Проекты"

    _write_projects_sheet(projects_sheet, submissions)
    _write_participants_sheet(workbook.create_sheet("Участники"), program, submissions)
    _write_evaluations_sheet(workbook.create_sheet("Оценки"), program)
    _write_totals_sheet(workbook.create_sheet("Итоги"), submissions)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _get_program_links(program: PartnerProgram):
    evaluations_qs = (
        ProjectEvaluation.objects.select_related("user")
        .prefetch_related(
            Prefetch(
                "evaluation_scores",
                queryset=ProjectEvaluationScore.objects.select_related(
                    "criterion"
                ).order_by("criterion_id"),
            )
        )
        .order_by("user_id")
    )

    return (
        PartnerProgramProject.objects.filter(partner_program=program)
        .select_related("partner_program", "project", "project__leader")
        .prefetch_related(
            Prefetch(
                "project__collaborator_set",
                queryset=Collaborator.objects.select_related("user").order_by("id"),
                to_attr="_analytics_collaborators",
            ),
            Prefetch(
                "project__links",
                queryset=ProjectLink.objects.order_by("id"),
                to_attr="_analytics_links",
            ),
            Prefetch(
                "evaluations", queryset=evaluations_qs, to_attr="_analytics_evaluations"
            ),
        )
        .order_by("-datetime_submitted", "-id")
    )


def _get_assignment_counts(
    program: PartnerProgram, links: list[PartnerProgramProject]
) -> dict[int, int]:
    if not program.is_distributed_evaluation:
        return {}

    project_ids = [program_project.project_id for program_project in links]
    if not project_ids:
        return {}

    rows = (
        ProjectExpertAssignment.objects.filter(
            partner_program=program,
            project_id__in=project_ids,
        )
        .values("project_id")
        .annotate(total=Count("id"))
    )
    return {row["project_id"]: row["total"] for row in rows}


def _get_default_required_evaluations(program: PartnerProgram) -> int:
    if program.max_project_rates is not None:
        return int(program.max_project_rates)
    return program.experts.count()


def _build_submission(
    *,
    program_project: PartnerProgramProject,
    assignment_counts: dict[int, int],
    default_required: int,
) -> dict:
    project = program_project.project
    participant_records = _get_project_participants(program_project)
    participant_names = [record["full_name"] for record in participant_records]
    participants_count = len(participant_records)
    submitted_evaluations = [
        evaluation
        for evaluation in getattr(program_project, "_analytics_evaluations", [])
        if evaluation.status == ProjectEvaluation.STATUS_SUBMITTED
    ]
    score_values = [
        Decimal(str(evaluation.total_score))
        for evaluation in submitted_evaluations
        if evaluation.total_score is not None
    ]
    evaluations_required = (
        assignment_counts.get(project.id, 0)
        if program_project.partner_program.is_distributed_evaluation
        else default_required
    )
    evaluations_received = len(submitted_evaluations)
    evaluation_status = _get_evaluation_status(
        evaluations_received=evaluations_received,
        evaluations_required=evaluations_required,
    )
    average_score = _average(score_values)
    author_name = _user_display_name(project.leader) or "Индивидуальный проект"
    participant_label = (
        author_name
        if program_project.partner_program.participation_format
        == PartnerProgram.PARTICIPATION_FORMAT_INDIVIDUAL
        else _format_participant_count(participants_count)
    )

    return {
        "program_project_id": program_project.id,
        "project_id": project.id,
        "project_title": project.name or "Проект без названия",
        "project_description": project.description or "",
        "participants_preview": _format_participants_preview(participant_names),
        "participants_count": participants_count,
        "author_name": author_name,
        "participant_label": participant_label,
        "submitted_at": program_project.datetime_submitted,
        "submission_status": (
            SUBMISSION_STATUS_SUBMITTED
            if program_project.submitted
            else SUBMISSION_STATUS_NOT_SUBMITTED
        ),
        "evaluation_status": evaluation_status,
        "evaluations_received": evaluations_received,
        "evaluations_required": evaluations_required,
        "average_score": _decimal_to_number(average_score),
        "final_score": _decimal_to_number(average_score),
        "project_materials": _get_project_materials(project),
    }


def _get_project_participants(program_project: PartnerProgramProject) -> list[dict]:
    project = program_project.project
    participants = []
    seen_user_ids = set()

    def add_user(user, role: str):
        if not user or user.id in seen_user_ids:
            return
        seen_user_ids.add(user.id)
        participants.append(
            {
                "user": user,
                "user_id": user.id,
                "full_name": _user_display_name(user),
                "email": user.email,
                "phone": getattr(user, "phone_number", "") or "",
                "role": role,
            }
        )

    add_user(project.leader, "Лидер")

    collaborators = getattr(project, "_analytics_collaborators", None)
    if collaborators is None:
        collaborators = project.collaborator_set.select_related("user").order_by("id")

    for collaborator in collaborators:
        add_user(collaborator.user, collaborator.role or "Участник")

    return participants


def _get_project_materials(project) -> list[dict]:
    materials = []
    if project.presentation_address:
        materials.append(
            {
                "title": "Презентация",
                "url": project.presentation_address,
                "kind": "presentation",
            }
        )

    links = getattr(project, "_analytics_links", None)
    if links is None:
        links = project.links.order_by("id")

    for index, link in enumerate(links, start=1):
        materials.append(
            {
                "title": f"Материал {index}",
                "url": link.link,
                "kind": "link",
            }
        )
    return materials


def _get_evaluation_status(
    *, evaluations_received: int, evaluations_required: int
) -> str:
    if evaluations_required <= 0 or evaluations_received <= 0:
        return EVALUATION_STATUS_NOT_EVALUATED
    if evaluations_received < evaluations_required:
        return EVALUATION_STATUS_PARTIALLY_EVALUATED
    return EVALUATION_STATUS_EVALUATED


def _get_registered_participants_count(program: PartnerProgram) -> int:
    return (
        PartnerProgramUserProfile.objects.filter(
            partner_program=program, user__isnull=False
        )
        .values("user_id")
        .distinct()
        .count()
    )


def _get_current_stage(program: PartnerProgram) -> str:
    current_time = timezone.now()
    submission_deadline = program.get_project_submission_deadline()

    if (
        program.datetime_registration_ends
        and current_time <= program.datetime_registration_ends
    ):
        return "Регистрация"
    if submission_deadline and current_time <= submission_deadline:
        return "Прием проектов"
    if (
        program.datetime_evaluation_ends
        and current_time <= program.datetime_evaluation_ends
    ):
        return "Экспертная оценка"
    return "Итоги"


def _user_display_name(user) -> str:
    if not user:
        return ""
    full_name = user.get_full_name().strip() if hasattr(user, "get_full_name") else ""
    return full_name or getattr(user, "email", "") or f"Пользователь {user.id}"


def _format_participant_count(count: int) -> str:
    if count % 10 == 1 and count % 100 != 11:
        suffix = "участник"
    elif count % 10 in (2, 3, 4) and count % 100 not in (12, 13, 14):
        suffix = "участника"
    else:
        suffix = "участников"
    return f"{count} {suffix}"


def _format_participants_preview(names: list[str]) -> str:
    visible_names = [name for name in names if name][:3]
    preview = ", ".join(visible_names)
    hidden_count = max(len(names) - len(visible_names), 0)
    if hidden_count:
        return f"{preview} +{hidden_count}" if preview else f"+{hidden_count}"
    return preview


def _average(values: list[Decimal]) -> Decimal | None:
    if not values:
        return None
    return (sum(values) / Decimal(len(values))).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )


def _decimal_to_number(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _write_projects_sheet(sheet, submissions: list[dict]) -> None:
    sheet.append(
        [
            "ID проекта",
            "Название проекта",
            "Автор / участники",
            "Количество участников",
            "Дата сдачи",
            "Статус сдачи",
            "Оценок получено",
            "Оценок требуется",
            "Средний балл",
            "Статус оценки",
            "Ссылка на проект",
        ]
    )

    for submission in submissions:
        _append_safe(
            sheet,
            [
                submission["project_id"],
                submission["project_title"],
                submission["participant_label"],
                submission["participants_count"],
                submission["submitted_at"],
                _submission_status_label(submission["submission_status"]),
                submission["evaluations_received"],
                submission["evaluations_required"],
                submission["average_score"],
                _evaluation_status_label(submission["evaluation_status"]),
                _project_url(submission["project_id"]),
            ],
        )


def _write_participants_sheet(
    sheet, program: PartnerProgram, submissions: list[dict]
) -> None:
    include_contacts = bool(getattr(program, "_privacy_include_contacts", False))
    header = [
        "ID участника",
        "ФИО",
        "Проект",
        "Роль в проекте",
        "Дата регистрации",
        "Статус участия",
        "Проект сдан / не сдан",
    ]
    if include_contacts:
        header[2:2] = ["Email", "Телефон"]
    sheet.append(header)

    membership = _build_project_membership_map(program)
    profiles = (
        PartnerProgramUserProfile.objects.filter(partner_program=program)
        .select_related("user", "project")
        .order_by("datetime_created", "id")
    )
    submission_by_project_id = {
        submission["project_id"]: submission for submission in submissions
    }

    for profile in profiles:
        user = profile.user
        if not user:
            continue
        membership_item = membership.get(user.id)
        project = membership_item["project"] if membership_item else profile.project
        submission = submission_by_project_id.get(project.id) if project else None
        row = [
            user.id,
            _user_display_name(user),
            project.name if project else "",
            membership_item["role"] if membership_item else "Участник",
            profile.datetime_created,
            "Зарегистрирован",
            (
                _submission_status_label(submission["submission_status"])
                if submission
                else "Не сдан"
            ),
        ]
        if include_contacts:
            row[2:2] = [
                user.email,
                getattr(user, "phone_number", "") or "",
            ]
        _append_safe(sheet, row)


def _write_evaluations_sheet(sheet, program: PartnerProgram) -> None:
    sheet.append(
        [
            "Проект",
            "Эксперт",
            "Статус оценки",
            "Дата отправки оценки",
            "Критерий",
            "Вес критерия",
            "Балл",
            "Комментарий эксперта",
        ]
    )

    evaluations = (
        ProjectEvaluation.objects.filter(
            program_project__partner_program=program,
            status=ProjectEvaluation.STATUS_SUBMITTED,
        )
        .select_related("user", "program_project__project")
        .prefetch_related(
            Prefetch(
                "evaluation_scores",
                queryset=ProjectEvaluationScore.objects.select_related(
                    "criterion"
                ).order_by("criterion_id"),
            )
        )
        .order_by("program_project__project__name", "user_id")
    )

    for evaluation in evaluations:
        scores = list(evaluation.evaluation_scores.all())
        if not scores:
            _append_safe(
                sheet,
                [
                    evaluation.program_project.project.name,
                    _user_display_name(evaluation.user),
                    "Оценено",
                    evaluation.submitted_at,
                    "",
                    "",
                    "",
                    evaluation.comment,
                ],
            )
            continue

        for score in scores:
            _append_safe(
                sheet,
                [
                    evaluation.program_project.project.name,
                    _user_display_name(evaluation.user),
                    "Оценено",
                    evaluation.submitted_at,
                    score.criterion.name,
                    score.criterion.weight,
                    score.value,
                    evaluation.comment,
                ],
            )


def _write_totals_sheet(sheet, submissions: list[dict]) -> None:
    sheet.append(
        [
            "Место",
            "Проект",
            "Автор / участники",
            "Средний балл",
            "Количество оценок",
            "Статус оценки",
            "Примечание",
        ]
    )

    ranked_submissions = sorted(
        submissions,
        key=lambda item: (
            item["average_score"] is not None,
            item["average_score"] or 0,
            item["evaluations_received"],
        ),
        reverse=True,
    )

    place = 1
    for submission in ranked_submissions:
        if submission["average_score"] is None:
            note = "Нет отправленных оценок"
            place_value = ""
        else:
            note = ""
            place_value = place
            place += 1

        _append_safe(
            sheet,
            [
                place_value,
                submission["project_title"],
                submission["participant_label"],
                submission["average_score"],
                submission["evaluations_received"],
                _evaluation_status_label(submission["evaluation_status"]),
                note,
            ],
        )


def _build_project_membership_map(program: PartnerProgram) -> dict[int, dict]:
    membership = {}
    links = _get_program_links(program)
    for program_project in links:
        for participant in _get_project_participants(program_project):
            membership[participant["user_id"]] = {
                "project": program_project.project,
                "role": participant["role"],
            }
    return membership


def _append_safe(sheet, values: list) -> None:
    sheet.append([sanitize_excel_value(value) for value in values])


def _project_url(project_id: int) -> str:
    return f"{settings.FRONTEND_URL.rstrip('/')}/office/projects/{project_id}"


def _submission_status_label(status: str) -> str:
    return "Сдан" if status == SUBMISSION_STATUS_SUBMITTED else "Не сдан"


def _evaluation_status_label(status: str) -> str:
    labels = {
        EVALUATION_STATUS_NOT_EVALUATED: "Не оценено",
        EVALUATION_STATUS_PARTIALLY_EVALUATED: "Частично",
        EVALUATION_STATUS_EVALUATED: "Оценено",
    }
    return labels.get(status, "Не оценено")
