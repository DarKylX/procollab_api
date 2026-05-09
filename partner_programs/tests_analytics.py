import io
from decimal import Decimal

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from partner_programs.models import (
    PartnerProgram,
    PartnerProgramProject,
    PartnerProgramUserProfile,
    PersonalDataAccessLog,
)
from projects.models import Collaborator, Project
from project_rates.models import Criteria, ProjectEvaluation, ProjectEvaluationScore
from users.models import CustomUser


class PartnerProgramAnalyticsTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.now = timezone.now()
        self.manager = self.create_user("manager-analytics@example.com")
        self.staff = self.create_user("staff-analytics@example.com", is_staff=True)
        self.outsider = self.create_user("outsider-analytics@example.com")
        self.leader = self.create_user(
            "leader-analytics@example.com", first_name="Leader"
        )
        self.collaborator = self.create_user(
            "collaborator-analytics@example.com",
            first_name="Collaborator",
        )
        self.expert_one = self.create_user(
            "expert-one-analytics@example.com",
            first_name="Expert",
            user_type=CustomUser.EXPERT,
        )
        self.expert_two = self.create_user(
            "expert-two-analytics@example.com",
            first_name="Second",
            user_type=CustomUser.EXPERT,
        )
        self.program = self.create_program()
        self.program.managers.add(self.manager)
        self.expert_one.expert.programs.add(self.program)
        self.expert_two.expert.programs.add(self.program)

        PartnerProgramUserProfile.objects.create(
            user=self.leader,
            partner_program=self.program,
            partner_program_data={},
        )
        PartnerProgramUserProfile.objects.create(
            user=self.collaborator,
            partner_program=self.program,
            partner_program_data={},
        )

        self.project = Project.objects.create(
            leader=self.leader,
            draft=False,
            is_public=False,
            name="Risk Model",
            description="Project description",
        )
        Collaborator.objects.create(
            project=self.project,
            user=self.collaborator,
            role="Analyst",
        )
        self.program_project = PartnerProgramProject.objects.create(
            partner_program=self.program,
            project=self.project,
            submitted=True,
            datetime_submitted=self.now,
        )
        self.criterion = Criteria.objects.create(
            name="Impact",
            description="Impact",
            type="float",
            min_value=1,
            max_value=10,
            weight=100,
            partner_program=self.program,
        )

    def create_user(self, email: str, **overrides):
        defaults = {
            "password": "pass",
            "first_name": "Test",
            "last_name": "User",
            "birthday": "1990-01-01",
            "is_active": True,
        }
        defaults.update(overrides)
        return CustomUser.objects.create_user(email=email, **defaults)

    def create_program(self, **overrides):
        defaults = {
            "name": "Financial Foresight 2026",
            "tag": "financial_foresight",
            "description": "Program description",
            "city": "Moscow",
            "data_schema": {},
            "status": "published",
            "draft": False,
            "is_competitive": True,
            "projects_availability": "all_users",
            "max_project_rates": 2,
            "datetime_registration_ends": self.now - timezone.timedelta(days=3),
            "datetime_project_submission_ends": self.now - timezone.timedelta(days=1),
            "datetime_evaluation_ends": self.now + timezone.timedelta(days=5),
            "datetime_started": self.now - timezone.timedelta(days=10),
            "datetime_finished": self.now + timezone.timedelta(days=20),
        }
        defaults.update(overrides)
        return PartnerProgram.objects.create(**defaults)

    def analytics_url(self):
        return f"/programs/{self.program.id}/analytics/"

    def export_url(self):
        return f"/programs/{self.program.id}/analytics/export/"

    def contact_export_url(self):
        return f"/programs/{self.program.id}/analytics/contact-export/"

    def create_submitted_evaluation(self, user, total_score: str):
        evaluation = ProjectEvaluation.objects.create(
            program_project=self.program_project,
            user=user,
            status=ProjectEvaluation.STATUS_SUBMITTED,
            total_score=Decimal(total_score),
            submitted_at=self.now,
            comment="Good",
        )
        ProjectEvaluationScore.objects.create(
            evaluation=evaluation,
            criterion=self.criterion,
            value=total_score,
        )
        return evaluation

    def test_manager_gets_analytics_with_project_participants(self):
        self.client.force_authenticate(self.manager)

        response = self.client.get(self.analytics_url())

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["participants_count"], 2)
        self.assertEqual(response.data["submitted_projects_count"], 1)
        submission = response.data["submissions"][0]
        self.assertEqual(submission["project_title"], "Risk Model")
        self.assertEqual(submission["participants_count"], 2)
        self.assertEqual(submission["participant_label"], "2 участника")
        self.assertIn("Leader", submission["participants_preview"])
        self.assertIn("Collaborator", submission["participants_preview"])
        self.assertNotIn("team_name", submission)

    def test_staff_gets_analytics_and_outsider_is_forbidden(self):
        self.client.force_authenticate(self.staff)
        staff_response = self.client.get(self.analytics_url())
        self.assertEqual(staff_response.status_code, 200)

        self.client.force_authenticate(self.outsider)
        outsider_response = self.client.get(self.analytics_url())
        self.assertEqual(outsider_response.status_code, 403)

        self.client.force_authenticate(self.leader)
        participant_response = self.client.get(self.analytics_url())
        self.assertEqual(participant_response.status_code, 403)

        self.client.force_authenticate(self.expert_one)
        expert_response = self.client.get(self.analytics_url())
        self.assertEqual(expert_response.status_code, 403)

    def test_manager_contact_export_requires_verified_program(self):
        self.client.force_authenticate(self.manager)

        analytics_response = self.client.get(self.analytics_url())
        contact_response = self.client.get(self.contact_export_url())

        self.assertEqual(analytics_response.status_code, 200)
        self.assertFalse(analytics_response.data["can_export_contacts"])
        self.assertEqual(contact_response.status_code, 403)

    def test_verified_manager_can_export_contacts(self):
        self.program.verification_status = PartnerProgram.VERIFICATION_STATUS_VERIFIED
        self.program.save(update_fields=["verification_status"])
        self.client.force_authenticate(self.manager)

        response = self.client.get(self.contact_export_url())

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        participants_sheet = workbook["Участники"]
        headers = [cell.value for cell in next(participants_sheet.iter_rows(max_row=1))]
        self.assertIn("Email", headers)
        self.assertIn("Телефон", headers)
        self.assertTrue(
            PersonalDataAccessLog.objects.filter(
                program=self.program,
                action=PersonalDataAccessLog.ACTION_PARTICIPANT_EXPORT_DOWNLOAD,
                metadata__export_type="analytics_contacts",
            ).exists()
        )

    def test_evaluation_required_zero_is_not_evaluated(self):
        self.program.max_project_rates = 0
        self.program.save(update_fields=["max_project_rates"])
        self.create_submitted_evaluation(self.expert_one, "9.00")
        self.client.force_authenticate(self.manager)

        response = self.client.get(self.analytics_url())

        self.assertEqual(response.status_code, 200)
        submission = response.data["submissions"][0]
        self.assertEqual(submission["evaluations_required"], 0)
        self.assertEqual(submission["evaluations_received"], 1)
        self.assertEqual(submission["evaluation_status"], "not_evaluated")
        self.assertEqual(response.data["evaluated_projects_count"], 0)

    def test_evaluation_status_and_average_score_are_recalculated(self):
        self.create_submitted_evaluation(self.expert_one, "8.00")
        self.client.force_authenticate(self.manager)

        partial_response = self.client.get(self.analytics_url())

        self.assertEqual(partial_response.status_code, 200)
        partial_submission = partial_response.data["submissions"][0]
        self.assertEqual(partial_submission["evaluation_status"], "partially_evaluated")
        self.assertEqual(partial_submission["average_score"], 8.0)
        self.assertEqual(partial_response.data["evaluated_projects_count"], 0)

        self.create_submitted_evaluation(self.expert_two, "10.00")

        evaluated_response = self.client.get(self.analytics_url())
        evaluated_submission = evaluated_response.data["submissions"][0]
        self.assertEqual(evaluated_submission["evaluation_status"], "evaluated")
        self.assertEqual(evaluated_submission["average_score"], 9.0)
        self.assertEqual(evaluated_response.data["evaluated_projects_count"], 1)
        self.assertEqual(evaluated_response.data["average_score"], 9.0)

    def test_analytics_export_contains_required_sheets(self):
        self.create_submitted_evaluation(self.expert_one, "8.00")
        self.client.force_authenticate(self.manager)

        response = self.client.get(self.export_url())

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Проекты", "Участники", "Оценки", "Итоги"],
        )
        projects_sheet = workbook["Проекты"]
        headers = [cell.value for cell in next(projects_sheet.iter_rows(max_row=1))]
        self.assertIn("Автор / участники", headers)
        self.assertNotIn("Команда / автор", headers)
